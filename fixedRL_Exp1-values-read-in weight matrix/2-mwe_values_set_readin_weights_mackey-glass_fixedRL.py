'''Exp 1 fixed RC - Mackey-Glass
The R is fixed, at each trial we create a new R, for which we then run n_trials.
This way we make sure that the final results of the n_trial do not depend on the one specific structure of the one R.
No zeros or close-to zero values are allowed in the read-in matrix.
No extra fraction input: fraction_input=1.0'''

import numpy as np
import os
import time
import argparse
import brainpy as bp
import brainpy.math as bm
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook, Workbook

from pyreco.custom_models import RC
from pyreco.layers import InputLayer, ReadoutLayer, RandomReservoirLayer
from pyreco.optimizers import RidgeSK

def create_base_model(input_shape, output_shape):
    model_rc = RC()
    model_rc.add(InputLayer(input_shape=input_shape))
    model_rc.add(RandomReservoirLayer(
        nodes=200, density=0.1, activation="tanh",
        leakage_rate=0.1, fraction_input=1.0, spec_rad=0.9 # no fraction_input
        ))
    model_rc.add(ReadoutLayer(output_shape, fraction_out=1.0))
    model_rc.compile(optimizer=RidgeSK(alpha=0.1), metrics=["mean_squared_error"])
    return model_rc

def generate_mackey_glass_data(n_samples=20000, beta=0.2, gamma=0.1, tau=17, n=10):
    class MackeyGlassEq(bp.Dynamic):
        def __init__(self, num):
            super().__init__(num)
            self.beta, self.gamma, self.tau, self.n = beta, gamma, tau, n
            self.delay_len = int(self.tau / bm.get_dt())
            self.x = bm.Variable(bm.zeros(num))
            self.x_delay = bm.LengthDelay(
                self.x, delay_len=self.delay_len,
                initial_delay_data=lambda sh, dtype: 1.2 + 0.2 * (bm.random.random(sh) - 0.5))
            self.x_oldest = bm.Variable(self.x_delay(self.delay_len))
            self.integral = bp.odeint(
                lambda x, t, x_tau: self.beta * x_tau / (1 + x_tau ** self.n) - self.gamma * x,
                method='exp_auto')

        def update(self):
            self.x.value = self.integral(self.x.value, bp.share['t'], self.x_oldest.value, bp.share['dt'])
            self.x_delay.update(self.x.value)
            self.x_oldest.value = self.x_delay(self.delay_len)

    runner = bp.DSRunner(MackeyGlassEq(1), monitors=['x', 'x_oldest'])
    runner.run(n_samples * bm.get_dt())
    data = np.column_stack([runner.mon.ts, runner.mon.x, runner.mon.x_oldest])
    return data


def mackey_glass_pred(n_batch, n_time_in, n_time_out, n_states=2):
    data = generate_mackey_glass_data(n_samples=n_batch + n_time_in + n_time_out)
    time_series = data[:, 1:1 + n_states]
    X = np.array([time_series[i:i + n_time_in] for i in range(n_batch)])
    y = np.array([time_series[i + n_time_in:i + n_time_in + n_time_out] for i in range(n_batch)])
    return train_test_split(X, y, test_size=0.2, random_state=42)


def sequence_to_scalar(n_batch=200, n_states=2):
    return mackey_glass_pred(n_batch=n_batch, n_time_in=1, n_time_out=1, n_states=n_states)


def sample_laplace_no_small(shape, threshold=1e-3, loc=0.0, scale=0.5):
    weights = np.random.laplace(loc=loc, scale=scale, size=shape)
    mask = np.abs(weights) < threshold
    while np.any(mask):
        weights[mask] = np.random.laplace(loc=loc, scale=scale, size=np.sum(mask))
        mask = np.abs(weights) < threshold
    return weights

def sample_fourier_weights(shape, max_freq=10, threshold=1e-3, max_attempts=100):
    n_nodes, n_inputs = shape
    t = 0.5
    
    freqs = np.random.uniform(0, max_freq, size=shape)
    phases = np.random.uniform(0, 2 * np.pi, size=shape)
    weights = np.sin(2 * np.pi * freqs * t + phases) + np.cos(2 * np.pi * freqs * t + phases)
    
    mask = np.abs(weights) < threshold
    attempts = 0
    
    while np.any(mask) and attempts < max_attempts:
        num_masked = np.count_nonzero(mask)
        freqs[mask] = np.random.uniform(0, max_freq, size=num_masked)
        phases[mask] = np.random.uniform(0, 2 * np.pi, size=num_masked)
        weights[mask] = np.sin(2 * np.pi * freqs[mask] * t + phases[mask]) + \
                        np.cos(2 * np.pi * freqs[mask] * t + phases[mask])
        mask = np.abs(weights) < threshold
        attempts += 1
    
    if np.any(mask):
        # Force remaining small weights to threshold (sign preserved)
        weights[mask] = np.sign(weights[mask]) * threshold
    
    # Normalize only once here:
    weights = 0.5 * weights / np.max(np.abs(weights))
    
    return weights

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--n_trials', type=int, default=100, help='Number of trials')
    args = parser.parse_args()

    bm.set_dt(0.05)

    excel_path = os.path.join(os.getcwd(), 'Losses_Read-in_FixedRL.xlsx')
    print(f"Excel path: {excel_path}")

    sheet_name_median = 'Exp1_Mackey-Glass_Median'
    sheet_name_iqr = 'Exp1_Mackey-Glass_IQR'

    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        wb = Workbook()

    ws_m = wb[sheet_name_median] if sheet_name_median in wb.sheetnames else wb.create_sheet(title=sheet_name_median)
    ws_iqr = wb[sheet_name_iqr] if sheet_name_iqr in wb.sheetnames else wb.create_sheet(title=sheet_name_iqr)

    for outer in range(args.n_trials):
        print(f"\n--- Outer Trial {outer + 1}/{args.n_trials} ---")
        X_train, X_test, y_train, y_test = sequence_to_scalar(n_batch=200, n_states=2)
        input_shape = (X_train.shape[1], X_train.shape[2])
        output_shape = (y_train.shape[1], y_train.shape[2])

        # Create RC model once per outer trial
        model_rc = create_base_model(input_shape, output_shape)

        losses = {'default': [], 'random': [], 'laplace': [], 'fourier': []}

        for inner in range(args.n_trials):
            print(f"  Inner Trial {inner + 1}/{args.n_trials}", end="\r")

            # Default input weights
            model_rc.fit(X_train, y_train)

            loss_default = model_rc.evaluate(X_test, y_test, metrics=["mae"])[0]
            losses['default'].append(loss_default)
            
            # Random input weights
            W_rand = np.random.randn(200, 2)
            while np.any(np.abs(W_rand) < 1e-3):
                W_rand[np.abs(W_rand) < 1e-3] =  np.random.randn(np.sum(np.abs(W_rand) < 1e-3))
            model_rc._set_readin_weights(W_rand)
            model_rc.fit(X_train, y_train)
            losses['random'].append(model_rc.evaluate(X_test, y_test, metrics=["mae"])[0])
            
            # Laplace input weights
            W_lap = sample_laplace_no_small((200, 2))
            model_rc._set_readin_weights(W_lap)
            model_rc.fit(X_train, y_train)
            losses['laplace'].append(model_rc.evaluate(X_test, y_test, metrics=["mae"])[0])
            
            # Fourier input weights
            W_fourier = sample_fourier_weights((200, 2))
            model_rc._set_readin_weights(W_fourier)
            model_rc.fit(X_train, y_train)
            losses['fourier'].append(model_rc.evaluate(X_test, y_test, metrics=["mae"])[0])
            
        medians = [np.median(losses[k]) for k in losses]
        iqrs = [np.percentile(losses[k], 75) - np.percentile(losses[k], 25) for k in losses]

        # Write medians to median sheet
        for i, val in enumerate(medians, start=1):
            ws_m.cell(row=outer + 1, column=i, value=val)

        # Write IQRs to IQR sheet
        for i, val in enumerate(iqrs, start=1):
            ws_iqr.cell(row=outer + 1, column=i, value=val)

    wb.save(excel_path)
    print(f"Results written to Excel row {outer + 1}")

if __name__ == "__main__":
    start = time.time()
    main()
    print(f"\nDone. Total time: {time.time() - start:.2f} seconds")
