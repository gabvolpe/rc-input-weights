'''Exp 1 fixed RC - Sine Wave
The R is fixed, at each trial we create a new R, for which we then run n_trials.
This way we make sure that the final results of the n_trial do not depend on the one specific structure of the one R.
No zeros or close-to zero values are allowed in the read-in matrix.
No extra fraction input: fraction_input=1.0'''

import os
import sys
import copy
import pickle
from matplotlib import pyplot as plt
import numpy as np
from pyreco.custom_models import RC
from pyreco.layers import InputLayer, ReadoutLayer, RandomReservoirLayer
from pyreco.utils_data import sequence_to_scalar
from pyreco.optimizers import RidgeSK
from pyreco.metrics import r2, mae, mse
import time
import argparse
from openpyxl import load_workbook, Workbook


# as its not a library, we need to add the parent directory to the path
parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if parent_dir not in sys.path:
    sys.path.append(parent_dir)

from helper_files.utils import create_weights
from helper_files.utils import compute_median_and_iqr_loss


def create_base_model(input_shape, output_shape):
    model = RC()
    model.add(InputLayer(input_shape=input_shape))
    model.add(RandomReservoirLayer(nodes=200, 
                                   density=0.1, 
                                   activation="tanh", 
                                   leakage_rate=0.1, 
                                   fraction_input=1.0)) # no fraction_input
    model.add(ReadoutLayer(output_shape, fraction_out=1.0)) #fraction_out either 1.0 or 0.5
    optim = RidgeSK(alpha=0.5)
    model.compile(optimizer=optim, metrics=["mean_squared_error"])
    return model




def main():
    parser = argparse.ArgumentParser(description="Run RC model with read-in weights variations.")
    parser.add_argument('--task', type=str, default='sine_prediction', help='Task name')
    parser.add_argument('--n_trials', type=int, default=10, help='Number of trials')
    args = parser.parse_args()

    if args.task == "sine_prediction":
        X_train, X_test, y_train, y_test = sequence_to_scalar(
            name="sine_prediction",
            n_states=1,
            n_batch=200,
            n_time_in=20,
        )
        input_shape = (X_train.shape[1], X_train.shape[2])
        output_shape = (y_train.shape[1], y_train.shape[2])
    else:
        raise NotImplementedError(f"Task {args.task} not implemented")
    
    # dict of ways to create the read-in weights (random normal, laplace, fourier, ...)
    weight_methods = {
        "RandomUniform": "random_uniform",
        "RandomNormal": "random_normal",
        "Laplace": "laplace",
        "Fourier": "fourier",
    }

    # dict of error metrics to track. Keys should match the model's compile metrics
    metrics = {"r2": r2, "mae": mae, "mse": mse}  

    # a bit of excel stuff
    excel_path = os.path.join(os.getcwd(), 'Losses_Read-in_FixedRL.xlsx')
    print(excel_path)
    sheet_name_median = 'Exp1_Sine Wave_Median'
    sheet_name_iqr = 'Exp1_Sine Wave_IQR'

    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        wb = Workbook()

    if sheet_name_median in wb.sheetnames:
        ws_m = wb[sheet_name_median]
    else:
        ws_m = wb.create_sheet(title=sheet_name_median)

    if sheet_name_iqr in wb.sheetnames:
        ws_iqr = wb[sheet_name_iqr]
    else:
        ws_iqr = wb.create_sheet(title=sheet_name_iqr)

    # # Prepare data structures to store losses: Each method will have a 2D list: [outer_trial][inner_trial]
    # losses = {
    #     "Old": [],
    #     "New": [],
    #     "Laplace": [],
    #     "Fourier": []
    # }
    # create loss dictionary for each readin weights method
    losses = {"outer_trial": np.zeros(args.n_trials**2)}  # let's keep the info about the outer trial
    losses["Baseline"] = np.zeros(args.n_trials**2)  # for the baseline model
    for method in weight_methods.keys():
        losses[method] = np.zeros((args.n_trials**2, len(metrics)))  # for each method, we store the losses for each metric


    start_time = time.time()
    i = 0 # iteration counter for the losses dict
    for trial_outer in range(args.n_trials):
        print(f"Outer Trial {trial_outer+1}/{args.n_trials} - Creating fresh model")
        
        # create a based RC model using a random reservoir layer
        model_rc = create_base_model(input_shape, output_shape)

        # inner trial: vary the read-in weights (keeping the reservoir fixed)
        for trial_inner in range(args.n_trials):
            # now we will vary the read-in weights
            print(f"\tInner Trial {trial_inner+1}/{args.n_trials}")

            # update outer trial index in the losses dict
            losses["outer_trial"][i] = trial_outer

            # create a clean copy for every inner trial
            _model = copy.deepcopy(model_rc)

            # obtain the baseline results (we only need this once, nothing 
            # changes here in the inner trials)
            if trial_inner == 0:
                _model.fit(X_train, y_train)
                _loss_bsl = _model.evaluate(X_test, y_test, 
                                            metrics=metrics.keys())
            # append loss to the losses dict based on the keys
            losses["Baseline"][i] = _loss_bsl[0]

            # now loop through the different read-in weights methods,
            # compute losses and store them
            for method, method_name in weight_methods.items():
                print(f"\t\tUsing {method_name} weights")
                readin_weights = create_weights((200, 1), method_name)
                _model._set_readin_weights(readin_weights)
                _model.fit(X_train, y_train)
                _loss = _model.evaluate(X_test, y_test, metrics=metrics.keys())
                losses[method][i] = _loss
                del _loss
            
            del _model
            i += 1

    # we should save the results to a pickle file
    with open('losses.pkl', 'wb') as f:
        pickle.dump(losses, f)

    """
    Post-processing the trials
    """

    # A: global summary of the losses per weight method
    for _method, _losses in losses.items():
        if _method == "outer_trial":
            continue

        # compute median and IQR for each method 
        # (across different reservoirs)
        for _idx_metric in range(len(metrics.items())):
            
            median, iqr = compute_median_and_iqr_loss(_losses[:, _idx_metric])

            metric = metrics.keys()[_idx_metric]
            print(f"{_method} - {metric}: Median = {median:.6f}, IQR = {iqr:.6f}")

    # B: summary per outer trial (i.e. for each reservoir)

     # create a summary post-processing dict (summarizing across inner trials)
    summary = {}
    summary["Baseline"] = {"median": [], "iqr": []}
    for method in weight_methods.keys():
        for metric in metrics.keys():
            summary[method][metric] = {"median": [], "iqr": []}


    for trial_outer in range(args.n_trials):
        idx_outer = losses["outer_trial"] == trial_outer

        for _method, _losses in losses.items():
            if _method == "outer_trial":
                continue

            for _idx_metric in range(len(metrics.items())):
                median, iqr = compute_median_and_iqr_loss(_losses[idx_outer, _idx_metric])

                _metric = metrics.keys()[_idx_metric]
                summary[_method][_metric]["median"].append(median)
                summary[_method][_metric]["iqr"].append(iqr)

   
    #         # Write to Excel
    #         col_median = 1 + list(weight_methods.keys()).index(_method) if _method != "Baseline" else 1
    #         col_iqr = col_median + len(weight_methods)

    # # compute median and IQR for each method
    # idx_outer = losses["outer_trial"] == trial_outer

    #     for 

    #     for method, method_name in weight_methods.items():
    #         for _idx_metric in range(len(metrics)):
    #             # compute median and IQR for the current method and metric
    #             median, iqr = compute_median_and_iqr_loss(losses[method][idx_outer, _idx_metric])
                
    #             # Write median and IQR to Excel
    #             col_median = 1 + list(weight_methods.keys()).index(method)
    #         median, iqr = compute_median_and_iqr_loss(losses[method][idx_outer])


   





    #     # lets make a simple figure of the historgram of the losses
    #     plt.figure(figsize=(10, 6))
    #     plt.axvline(losses["Old"][trial_outer][0], label='Old Model')
    #     plt.hist(losses["New"][trial_outer], bins=30, alpha=0.5, label='Random Normal')
    #     plt.hist(losses["Laplace"][trial_outer], bins=30, alpha=0.5, label='Laplace Model')
    #     plt.hist(losses["Fourier"][trial_outer], bins=30, alpha=0.5, label='Fourier Model')
    #     plt.title(f'Loss Distribution for Outer Trial {trial_outer + 1}')
    #     plt.xlabel('Loss (MAE)')
    #     plt.ylabel('Frequency')
    #     plt.legend()
    #     plt.show()

    #     # Row in Excel for this outer trial
    #     summary_row = 1 + trial_outer  # Assuming no header, otherwise write 2 instead of 1

    #     # Define model name to column index mapping
    #     model_columns = {
    #         "Old": 1,
    #         "New": 2,
    #         "Laplace": 3,
    #         "Fourier": 4
    #     }

    #     print(f"\n==== Outer Trial {trial_outer + 1} Summary ====")

    #     for method, col in model_columns.items():
    #         all_losses_flat = losses[method][trial_outer]  # losses for this outer trial
    #         median, iqr = compute_median_and_iqr_loss(all_losses_flat)

    #         # Write median and IQR for this outer trial to Excel
    #         ws_m.cell(row=summary_row, column=col, value=median)
    #         ws_iqr.cell(row=summary_row, column=col, value=iqr)

    #         # Print the summary
    #         print(f"{method} Model — Median MAE: {median:.6f}, IQR: {iqr:.6f}")
            
    #         wb.save(excel_path)
    # end_time = time.time()
    print(f"\nTotal time taken: {end_time - start_time:.2f} seconds")

if __name__ == "__main__":
    main()
