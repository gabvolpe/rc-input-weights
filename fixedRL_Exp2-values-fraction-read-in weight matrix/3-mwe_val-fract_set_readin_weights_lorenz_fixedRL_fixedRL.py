'''Exp 2 fixed RC - Lorenz
The R is fixed, at each trial we create a new R, for which we then run n_trials.
This way we make sure that the final results of the n_trial do not depend on the one specific structure of the one R.
No zeros or close-to zero values are allowed in the read-in matrix.
Some values get masked according to a fraction input: fraction_input=0.5'''

import os
import numpy as np
import time
import argparse
from openpyxl import load_workbook, Workbook

from pyreco.custom_models import RC
from pyreco.layers import InputLayer, ReadoutLayer, RandomReservoirLayer
from pyreco.optimizers import RidgeSK

from sklearn.model_selection import train_test_split
from scipy.integrate import solve_ivp

def create_base_model(input_shape,output_shape):
    model = RC()
    model.add(InputLayer(input_shape=input_shape))
    model.add(RandomReservoirLayer(
        nodes=500,
        density=0.1,
        activation="tanh",
        leakage_rate=0.8,
        fraction_input=0.5, # 50% of values masked
        spec_rad=0.9,
    ))
    model.add(ReadoutLayer(output_shape, fraction_out=1.0))
    optim = RidgeSK(alpha=0.1)
    model.compile(optimizer=optim, metrics=["mean_squared_error"])
    return model

sigma, beta, rho = 10, 8/3, 28

def lorenz(t, state):
    x, y, z = state
    dx = sigma * (y - x)
    dy = x * (rho - z) - y
    dz = x * y - beta * z
    return [dx, dy, dz]

def generate_lorenz_data(n_samples=20000, t_span=(0, 10)):
    t_eval = np.linspace(t_span[0], t_span[1], n_samples)
    initial_state = [1.0, 1.0, 1.0]
    sol = solve_ivp(lorenz, t_span, initial_state, t_eval=t_eval)
    return sol.y.T

def lorenz_pred(n_batch, n_time_in, n_time_out, n_states):
    data = generate_lorenz_data(n_samples=n_batch + n_time_in + n_time_out)
    X, y = [], []
    for i in range(n_batch):
        X.append(data[i:i + n_time_in, :n_states])
        y.append(data[i + n_time_in:i + n_time_in + n_time_out, :n_states])
    return train_test_split(np.array(X), np.array(y), test_size=0.2, random_state=42)

def vector_to_vector(name, n_batch=500, n_states=3):
    n_time_in = 1
    n_batch = max(n_batch, 2)
    if name == "vector2vector":
        X_train, X_test, y_train, y_test = lorenz_pred(n_batch, n_time_in, 1, n_states)
    else:
        raise ValueError(f"Unsupported task {name}")
    print(f"Input shape: {X_train.shape}, Output shape: {y_train.shape}")
    return X_train, X_test, y_train, y_test

def sequence_to_sequence(name, n_batch=500, n_states=3, n_time_in=100, n_time_out=5):
    n_batch = max(n_batch, 2)
    if name == "sequence2sequence":
        X_train, X_test, y_train, y_test = lorenz_pred(n_batch, n_time_in, n_time_out, n_states)
    else:
        raise ValueError(f"Unsupported task {name}")
    print(f"Input shape: {X_train.shape}, Output shape: {y_train.shape}")
    return X_train, X_test, y_train, y_test

def sample_laplace_no_small(shape, threshold=1e-3, loc=0.0, scale=0.5):
    weights = np.random.laplace(loc=loc, scale=scale, size=shape)
    mask = np.abs(weights) < threshold
    while np.any(mask):
        weights[mask] = np.random.laplace(loc=loc, scale=scale, size=np.sum(mask))
        mask = np.abs(weights) < threshold
    return weights

def generate_fourier_weights(n_nodes, n_inputs, max_freq=10, threshold=1e-3):
    frequencies = np.random.uniform(0, max_freq, size=(n_nodes, n_inputs))
    phases = np.random.uniform(0, 2 * np.pi, size=(n_nodes, n_inputs))
    t = 0  # fixed time step
    input_weights = (np.sin(2 * np.pi * frequencies * t + phases) + 
                     np.cos(2 * np.pi * frequencies * t + phases))
    input_fourier_weights = 0.5 * input_weights / np.max(np.abs(input_weights))
    mask_small = np.abs(input_fourier_weights) < threshold
    while np.any(mask_small):
        for i in range(n_nodes):
            for j in range(n_inputs):
                if mask_small[i, j]:
                    frequencies[i, j] = np.random.uniform(0, max_freq)
                    phases[i, j] = np.random.uniform(0, 2 * np.pi)
                    input_weights[i, j] = (np.sin(2 * np.pi * frequencies[i, j] * t + phases[i, j]) +
                                          np.cos(2 * np.pi * frequencies[i, j] * t + phases[i, j]))
        input_fourier_weights = 0.5 * input_weights / np.max(np.abs(input_weights))
        mask_small = np.abs(input_fourier_weights) < threshold
    return input_fourier_weights


# --- Main program ---

def main():
    import os

    parser = argparse.ArgumentParser(description="Test CustomModel API on Lorenz data")
    parser.add_argument("--task", default="sequence2sequence", choices=["sequence2sequence", "vector2vector"])
    parser.add_argument("--n_trials", default=30, type=int, help="Number of outer trials")
    parser.add_argument("--n_inner_trials", default=30, type=int, help="Number of inner trials")
    args = parser.parse_args()

    excel_path = os.path.join(os.getcwd(), 'Losses_Read-in_FixedRL.xlsx')
    print(excel_path)

    sheet_name_median = 'Exp2_Lorenz_Median'
    sheet_name_iqr = 'Exp2_Lorenz_IQR'

    # Load or create workbook and sheets
    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        wb = Workbook()

    if sheet_name_median in wb.sheetnames:
        ws_median = wb[sheet_name_median]
    else:
        ws_median = wb.create_sheet(sheet_name_median)

    if sheet_name_iqr in wb.sheetnames:
        ws_iqr = wb[sheet_name_iqr]
    else:
        ws_iqr = wb.create_sheet(sheet_name_iqr)

    # Load data depending on task
    if args.task == "sequence2sequence":
        X_train, X_test, y_train, y_test = sequence_to_sequence("sequence2sequence", n_batch=500, n_states=3)
    else:
        X_train, X_test, y_train, y_test = vector_to_vector("vector2vector", n_batch=500, n_states=3)

    input_shape = (X_train.shape[1], X_train.shape[2])
    output_shape = (y_train.shape[1], y_train.shape[2])

    print(f"Input shape: {input_shape}, Output shape: {output_shape}")

    n_trials = args.n_trials
    n_inner_trials = args.n_inner_trials

    start_time = time.time()

    for outer in range(n_trials):
        print(f"\n--- Outer Trial {outer + 1}/{args.n_trials} ---")
        # Create fresh model for each outer trial
        model = create_base_model(input_shape, output_shape)

        # Store losses per inner trial
        losses_old = []
        losses_new = []
        losses_laplace = []
        losses_fourier = []

        for inner in range(n_inner_trials):
            print(f"  Inner Trial {inner + 1}/{args.n_trials}", end="\r")
            model.fit(X_train, y_train)
            print("     old model fit.")

            # Evaluate old weights loss
            loss_old = model.evaluate(X_test, y_test, metrics=["mae"])
            losses_old.append(loss_old[0])

            # New input weights (random normal, no fraction_input)
            new_input_weights = np.random.randn(500, 3)
            threshold = 1e-3
            max_attempts = 10
            attempt = 0
            while np.any(np.abs(new_input_weights) < threshold) and attempt<max_attempts:
                mask = np.abs(new_input_weights) < threshold
                new_input_weights[mask] = np.random.randn(np.sum(mask))
                attempt +=1

            model._set_readin_weights(new_input_weights)
            model.fit(X_train, y_train)
            loss_new = model.evaluate(X_test, y_test, metrics=["mae"])
            losses_new.append(loss_new[0])
            print("     new random model fit.")

            # Laplace weights
            input_laplace_weights = sample_laplace_no_small((500, 3))
            model._set_readin_weights(input_laplace_weights)
            model.fit(X_train, y_train)
            loss_laplace = model.evaluate(X_test, y_test, metrics=["mae"])
            losses_laplace.append(loss_laplace[0])
            print("     laplace model fit.")

            # Fourier weights
            input_fourier_weights = generate_fourier_weights(500, 3)
            model._set_readin_weights(input_fourier_weights)
            model.fit(X_train, y_train)
            loss_fourier = model.evaluate(X_test, y_test, metrics=["mae"])
            losses_fourier.append(loss_fourier[0])
            print("     fourier model fit.")

        # Compute median and IQR for each input weights type
        def compute_median_and_iqr(losses_list):
            median = np.median(losses_list)
            iqr = np.subtract(*np.percentile(losses_list, [75, 25]))
            return median, iqr

        # Assuming losses dict has keys 'Old', 'New', 'Laplace', 'Fourier'
        # Each key maps to a list of losses per outer trial (list of lists)

        # For the current outer trial index trial_outer (or outer in your code)
        summary_row = 1 + outer  # or 2 if you want to start from row 2 for header

        model_columns = {
            "Old": 1,
            "New": 2,
            "Laplace": 3,
            "Fourier": 4
        }

        print(f"\n==== Outer Trial {outer + 1} Summary ====")

        for method, col in model_columns.items():
            # Get losses list for this outer trial and method
            losses_list = {
                "Old": losses_old,
                "New": losses_new,
                "Laplace": losses_laplace,
                "Fourier": losses_fourier
            }[method]

            median, iqr = compute_median_and_iqr(losses_list)

            # Write median and IQR to Excel sheets
            ws_median.cell(row=summary_row, column=col, value=median)
            ws_iqr.cell(row=summary_row, column=col, value=iqr)

            print(f"{method} Model — Median MAE: {median:.6f}, IQR: {iqr:.6f}")

        wb.save(excel_path)

    total_time = time.time() - start_time
    print(f"Completed {n_trials} outer trials with {n_trials} inner trials each in {total_time:.2f} seconds.")


if __name__ == "__main__":
    main()