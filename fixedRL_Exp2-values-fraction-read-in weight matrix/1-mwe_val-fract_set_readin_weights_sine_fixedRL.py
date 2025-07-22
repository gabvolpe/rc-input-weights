'''Exp 2 fixed RC - Sine Wave
The R is fixed, at each trial we create a new R, for which we then run n_trials.
This way we make sure that the final results of the n_trial do not depend on the one specific structure of the one R.
No zeros or close-to zero values are allowed in the read-in matrix.
Some values get masked according to a fraction input: fraction_input=0.5'''

import os
import numpy as np
from pyreco.custom_models import RC
from pyreco.layers import InputLayer, ReadoutLayer, RandomReservoirLayer
from pyreco.utils_data import sequence_to_scalar
from pyreco.optimizers import RidgeSK
import time
import argparse
from openpyxl import load_workbook, Workbook

def create_base_model(input_shape, output_shape):
    model = RC()
    model.add(InputLayer(input_shape=input_shape))
    model.add(RandomReservoirLayer(nodes=200, density=0.1, activation="tanh", leakage_rate=0.1, fraction_input=0.5)) # 50% of values masked
    model.add(ReadoutLayer(output_shape, fraction_out=1.0)) #fraction_out either 1.0 or 0.5
    optim = RidgeSK(alpha=0.5)
    model.compile(optimizer=optim, metrics=["mean_squared_error"])
    return model

def create_weights(shape, method):
    threshold = 1e-3
    if method == "random_normal":
        w = np.random.randn(*shape)
        while np.any(np.abs(w) < threshold):
            idx = np.abs(w) < threshold
            w[idx] = np.random.randn(np.sum(idx))
        return w

    elif method == "laplace":
        w = np.random.laplace(loc=0.0, scale=0.5, size=shape).flatten()
        while np.any(np.abs(w) < threshold):
            idx = np.abs(w) < threshold
            w[idx] = np.random.laplace(loc=0.0, scale=0.5, size=np.sum(idx))
        return w.reshape(shape)

    elif method == "fourier":
        num_weights = shape[0]
        max_freq = 10
        t = np.linspace(0, 1, num_weights).reshape(-1, 1)
        frequencies = np.random.uniform(0, max_freq, size=(num_weights, 1))
        phases = np.random.uniform(0, 2*np.pi, size=(num_weights, 1))
        weights = np.sin(2*np.pi*frequencies*t + phases) + np.cos(2*np.pi*frequencies*t + phases)
        weights = 0.5 * weights / np.max(np.abs(weights))
        weights = weights.flatten()

        # Fix: use np.where to get indices where weights are small
        while np.any(np.abs(weights) < threshold):
            idxs = np.where(np.abs(weights) < threshold)[0]

            # Regenerate frequencies and phases for these indices only
            frequencies[idxs, 0] = np.random.uniform(0, max_freq, size=len(idxs))
            phases[idxs, 0] = np.random.uniform(0, 2*np.pi, size=len(idxs))

            # Recompute weights at these indices
            # Since t is shape (num_weights, 1), we can index t[idxs, 0] to get scalar t values for each idx
            weights_temp = (
                np.sin(2*np.pi*frequencies[idxs, 0] * t[idxs, 0] + phases[idxs, 0]) +
                np.cos(2*np.pi*frequencies[idxs, 0] * t[idxs, 0] + phases[idxs, 0])
            )
            # Normalize partial weights by max absolute weight from full weights array
            max_abs = np.max(np.abs(weights))
            weights[idxs] = 0.5 * weights_temp / max_abs

        return weights.reshape(shape)

    else:
        raise ValueError(f"Unknown weight initialization method: {method}")

def compute_median_and_iqr(losses):
    numeric_losses = np.array([x for x in losses if isinstance(x, (int, float, np.float32, np.float64))])
    if len(numeric_losses) == 0:
        return None, None
    median = np.median(numeric_losses)
    q75, q25 = np.percentile(numeric_losses, [75, 25])
    iqr = q75 - q25
    return median, iqr

def main():
    parser = argparse.ArgumentParser(description="Run RC model with read-in weights variations.")
    parser.add_argument('--task', type=str, default='sine_prediction', help='Task name')
    parser.add_argument('--n_trials', type=int, default=100, help='Number of trials')
    args = parser.parse_args()

    if args.task == "sine_prediction":
        X_train, X_test, y_train, y_test = sequence_to_scalar(
            name="sine_prediction",
            n_states=1,
            n_batch=200,
            n_time_in=20,
        )
    else:
        raise NotImplementedError(f"Task {args.task} not implemented")

    input_shape = (X_train.shape[1], X_train.shape[2])
    output_shape = (y_train.shape[1], y_train.shape[2])

    excel_path = os.path.join(os.getcwd(), 'Losses_Read-in_FixedRL.xlsx')
    print(excel_path)
    sheet_name_median = 'Exp2_Sine Wave_Median'
    sheet_name_iqr = 'Exp2_Sine Wave_IQR'

    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        wb = Workbook()

    if sheet_name_median in wb.sheetnames:
        ws_m = wb[sheet_name_median]
    else:
        ws_m = wb.create_sheet(title=sheet_name_median)

    if sheet_name_iqr in wb.sheetnames:
        ws_iqr = wb[sheet_name_iqr]
    else:
        ws_iqr = wb.create_sheet(title=sheet_name_iqr)

    # Prepare data structures to store losses: Each method will have a 2D list: [outer_trial][inner_trial]
    losses = {
        "Old": [],
        "New": [],
        "Laplace": [],
        "Fourier": []
    }

    start_time = time.time()

    for trial_outer in range(args.n_trials):
        print(f"Outer Trial {trial_outer+1}/{args.n_trials} - Creating fresh model")
        model_rc = create_base_model(input_shape, output_shape)

        # Store losses for inner trials in this outer trial
        losses["Old"].append([])
        losses["New"].append([])
        losses["Laplace"].append([])
        losses["Fourier"].append([])

        for trial_inner in range(args.n_trials):
            print(f"  Inner Trial {trial_inner+1}/{args.n_trials}")

            # 1) Old: use model as is (initial weights)
            model_rc.fit(X_train, y_train)
            loss_old = model_rc.evaluate(X_test, y_test, metrics=["mae"])[0]
            losses["Old"][trial_outer].append(loss_old)

            # 2) New: random normal weights
            new_weights = create_weights((200,1), "random_normal")
            model_rc._set_readin_weights(new_weights)
            model_rc.fit(X_train, y_train)
            loss_new = model_rc.evaluate(X_test, y_test, metrics=["mae"])[0]
            losses["New"][trial_outer].append(loss_new)

            # 3) Laplace weights
            laplace_weights = create_weights((200,1), "laplace")
            model_rc._set_readin_weights(laplace_weights)
            model_rc.fit(X_train, y_train)
            loss_laplace = model_rc.evaluate(X_test, y_test, metrics=["mae"])[0]
            losses["Laplace"][trial_outer].append(loss_laplace)

            # 4) Fourier weights
            fourier_weights = create_weights((200,1), "fourier")
            model_rc._set_readin_weights(fourier_weights)
            model_rc.fit(X_train, y_train)
            loss_fourier = model_rc.evaluate(X_test, y_test, metrics=["mae"])[0]
            losses["Fourier"][trial_outer].append(loss_fourier)
    
        # Row in Excel for this outer trial
        summary_row = 1 + trial_outer  # Assuming no header, otherwise write 2 instead of 1

        # Define model name to column index mapping
        model_columns = {
            "Old": 1,
            "New": 2,
            "Laplace": 3,
            "Fourier": 4
        }

        print(f"\n==== Outer Trial {trial_outer + 1} Summary ====")

        for method, col in model_columns.items():
            all_losses_flat = losses[method][trial_outer]  # losses for this outer trial
            median, iqr = compute_median_and_iqr(all_losses_flat)

            # Write median and IQR for this outer trial to Excel
            ws_m.cell(row=summary_row, column=col, value=median)
            ws_iqr.cell(row=summary_row, column=col, value=iqr)

            # Print the summary
            print(f"{method} Model — Median MAE: {median:.6f}, IQR: {iqr:.6f}")
            
            wb.save(excel_path)
    end_time = time.time()
    print(f"\nTotal time taken: {end_time - start_time:.2f} seconds")

if __name__ == "__main__":
    main()
