
'''
Mackey-Glass
'''

import os
import numpy as np
import time
import argparse
from openpyxl import load_workbook, Workbook
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from pyreco.custom_models import RC
from pyreco.layers import InputLayer, ReadoutLayer, RandomReservoirLayer
from pyreco.optimizers import RidgeSK
from sklearn.model_selection import train_test_split
import brainpy as bp
import brainpy.math as bm
import concurrent.futures
import pickle
import networkx as nx
import random

# ----------------------
# --- User Arguments ---
# ----------------------
parser = argparse.ArgumentParser(description="Run RC model with customizable hyperparameters.")

parser.add_argument("--n_trials", type=int, default=100, help="Number of outer trials")
parser.add_argument("--n_inner", type=int, default=50, help="Number of inner trials")

parser.add_argument("--reservoir_nodes", type=int, default=200)
parser.add_argument("--density", type=float, default=0.2)
parser.add_argument("--spectral_radius", type=float, default=0.8)
parser.add_argument("--leakage_rate", type=float, default=0.5)
parser.add_argument("--fraction_input", type=float, default=1.0, help="Percentage of input without masking")

parser.add_argument("--ridge_alpha", type=float, default=0.1)

parser.add_argument("--set_threshold", type=bool, default=True, help="Set to False if no threshold wished")
parser.add_argument("--readin_threshold", type=float, default=1e-3)

parser.add_argument(
    "--sd_list",
    type=str,
    default="[0.1, 0.25, 0.5, 0.75, 1.0]",
    help="List of Gaussian SD values, format: [0.1,0.25,...]"
)
parser.add_argument('--task', type=str, default='sequence_to_sequence')

parser.add_argument(
    "--constraint_set", 
    type=str, default="1", 
    choices=["1", "2", "3"], 
    help="It saves the result into a specific Excel map according to the Constraint set")

# Parse
args = parser.parse_args()

# Convert SD list string → list
sd_list = eval(args.sd_list)

# ----------------------
# --- Functions --------
# ----------------------
def create_base_model(input_shape, output_shape):
    model_rc = RC()
    model_rc.add(InputLayer(input_shape=input_shape))
    reservoir_layer = RandomReservoirLayer(
        nodes=args.reservoir_nodes,
        density=args.density,
        activation="tanh",
        spec_rad=args.spectral_radius,
        leakage_rate=args.leakage_rate,
        fraction_input=args.fraction_input
    )
    model_rc.add(reservoir_layer)
    model_rc.add(ReadoutLayer(output_shape, fraction_out=1.0))
    optim = RidgeSK(alpha=args.ridge_alpha)
    model_rc.compile(optimizer=RidgeSK(alpha=0.1), metrics=["mean_squared_error"])
    return model_rc, reservoir_layer

def generate_mackey_glass_data(n_samples=20000, beta=0.2, gamma=0.1, tau=17, n=10):
    class MackeyGlassEq(bp.Dynamic):
        def __init__(self, num):
            super().__init__(num)
            self.beta, self.gamma, self.tau, self.n = beta, gamma, tau, n
            self.delay_len = int(self.tau / bm.get_dt())
            self.x = bm.Variable(bm.zeros(num))
            self.x_delay = bm.LengthDelay(
                self.x, delay_len=self.delay_len,
                initial_delay_data=lambda sh, dtype: 1.2 + 0.2 * (bm.random.random(sh) - 0.5))
            self.x_oldest = bm.Variable(self.x_delay(self.delay_len))
            self.integral = bp.odeint(
                lambda x, t, x_tau: self.beta * x_tau / (1 + x_tau ** self.n) - self.gamma * x,
                method='exp_auto')
        def update(self):
            self.x.value = self.integral(self.x.value, bp.share['t'], self.x_oldest.value, bp.share['dt'])
            self.x_delay.update(self.x.value)
            self.x_oldest.value = self.x_delay(self.delay_len)
    runner = bp.DSRunner(MackeyGlassEq(1), monitors=['x', 'x_oldest'])
    runner.run(n_samples * bm.get_dt())
    data = np.column_stack([runner.mon.ts, runner.mon.x, runner.mon.x_oldest])
    return data

def mackey_glass_pred(n_batch, n_time_in, n_time_out, n_states=2):
    data = generate_mackey_glass_data(n_samples=n_batch + n_time_in + n_time_out)
    time_series = data[:, 1:1 + n_states]
    X = np.array([time_series[i:i + n_time_in] for i in range(n_batch)])
    y = np.array([time_series[i + n_time_in:i + n_time_in + n_time_out] for i in range(n_batch)])
    return train_test_split(X, y, test_size=0.2, random_state=42)

# Sequence-to-Sequence
def sequence_to_sequence(n_batch=200, n_states=2):
    return mackey_glass_pred(n_batch=n_batch, n_states=n_states, n_time_in=1000, n_time_out=100)

task_functions = {
    "sequence_to_sequence": sequence_to_sequence
}

# Map string to function
args.task = task_functions[args.task]

# --- Create Weights -----

def create_weights(shape, method, dynamic_sd=None):
    if args.set_threshold is True:
        threshold = args.readin_threshold
        if method == "random_uniform":
            return np.random.uniform(-1, 1, size=shape)
        if method == "random_normal":
            mu = 0.0
            sd = dynamic_sd if dynamic_sd is not None else 1.0
            w = np.random.normal(mu, sd, size=shape)
            while np.any(np.abs(w) < threshold):
                idx = np.abs(w) < threshold
                w[idx] = np.random.normal(mu, sd, size=np.sum(idx))
            return w
        if method == "double_gaussian":
            mu1, mu2 = -1.5, 1.5
            sigma1, sigma2 = dynamic_sd, dynamic_sd
            amp1, amp2 = 0.5, 0.5
            n_elements = np.prod(shape)
            choices = np.random.choice([0, 1], size=n_elements, p=[amp1, amp2])
            g1 = np.random.normal(mu1, sigma1, size=n_elements)
            g2 = np.random.normal(mu2, sigma2, size=n_elements)
            w = np.where(choices == 0, g1, g2)
            while np.any(np.abs(w) < threshold):
                idx = np.abs(w) < threshold
                n_idx = np.sum(idx)
                g1_new = np.random.normal(mu1, sigma1, size=n_idx)
                g2_new = np.random.normal(mu2, sigma2, size=n_idx)
                w[idx] = np.where(np.random.rand(n_idx) < amp1, g1_new, g2_new)
            return w.reshape(shape)
        if method == "laplace":
            w = np.random.laplace(loc=0.0, scale=0.5, size=shape).flatten()
            while np.any(np.abs(w) < threshold):
                idx = np.abs(w) < threshold
                w[idx] = np.random.laplace(loc=0.0, scale=0.5, size=np.sum(idx))
            return w.reshape(shape)
        if method == "power_law":
            a = 2.0
            positive = np.random.power(a, size=np.prod(shape))
            negative = -positive.copy()
            combined = np.concatenate([positive, negative])
            w = np.random.choice(combined, size=np.prod(shape), replace=False)
            while np.any(np.abs(w) < threshold):
                idx = np.abs(w) < threshold
                n_new = np.sum(idx)
                new_pos = np.random.power(a, size=n_new // 2 + n_new % 2)
                new_neg = -np.random.power(a, size=n_new // 2)
                new_vals = np.concatenate([new_pos, new_neg])
                np.random.shuffle(new_vals)
                w[idx] = new_vals[:n_new]
            return w.reshape(shape)
        raise ValueError(f"Unknown weight initialization method: {method}")
    else:
        if method == "random_uniform":
            return np.random.uniform(-1, 1, size=shape)
        if method == "random_normal":
            mu = 0.0
            sd = dynamic_sd if dynamic_sd is not None else 1.0
            w = np.random.normal(mu, sd, size=shape)
            return w
        if method == "double_gaussian":
            mu1, mu2 = -1.5, 1.5
            sigma1, sigma2 = dynamic_sd, dynamic_sd
            amp1, amp2 = 0.5, 0.5
            n_elements = np.prod(shape)
            choices = np.random.choice([0, 1], size=n_elements, p=[amp1, amp2])
            g1 = np.random.normal(mu1, sigma1, size=n_elements)
            g2 = np.random.normal(mu2, sigma2, size=n_elements)
            w = np.where(choices == 0, g1, g2)
            return w.reshape(shape)
        if method == "laplace":
            w = np.random.laplace(loc=0.0, scale=0.5, size=shape).flatten()
            return w.reshape(shape)
        if method == "power_law":
            a = 2.0
            positive = np.random.power(a, size=np.prod(shape))
            negative = -positive.copy()
            combined = np.concatenate([positive, negative])
            w = np.random.choice(combined, size=np.prod(shape), replace=False)
            return w.reshape(shape)
        raise ValueError(f"Unknown weight initialization method: {method}")

def compute_median_and_iqr(losses):
    numeric_losses = np.array([x for x in losses if isinstance(x, (int, float, np.float32, np.float64))])
    if len(numeric_losses) == 0:
        return None, None
    median = np.median(numeric_losses)
    q75, q25 = np.percentile(numeric_losses, [75, 25])
    iqr = q75 - q25
    return median, iqr


def run_inner_trial(model_serialized, sd_list, X_train, y_train, X_test, y_test):
    model_rc = pickle.loads(model_serialized)
    trial_losses = {}
    trial_predictions = {}

    weights_uniform = create_weights((200,2), "random_uniform")
    model_rc._set_readin_weights(weights_uniform)
    model_rc.fit(X_train, y_train)
    y_pred = model_rc.predict(X_test)
    trial_predictions["Uniform"] = y_pred
    trial_losses["Uniform"] = model_rc.evaluate(X_test, y_test, metrics=["mae"])[0]

    for sd in sd_list:
        weights_gauss = create_weights((200,2), "random_normal", dynamic_sd=sd)
        model_rc._set_readin_weights(weights_gauss)
        model_rc.fit(X_train, y_train)
        y_pred = model_rc.predict(X_test)
        trial_predictions[f"Gaussian_sd_{sd}"] = y_pred
        trial_losses[f"Gaussian_sd_{sd}"] = model_rc.evaluate(X_test, y_test, metrics=["mae"])[0]

    weights_laplace = create_weights((200,2), "laplace")
    model_rc._set_readin_weights(weights_laplace)
    model_rc.fit(X_train, y_train)
    y_pred = model_rc.predict(X_test)
    trial_predictions["Laplace"] = y_pred
    trial_losses["Laplace"] = model_rc.evaluate(X_test, y_test, metrics=["mae"])[0]

    weights_power = create_weights((200,2), "power_law")
    model_rc._set_readin_weights(weights_power)
    model_rc.fit(X_train, y_train)
    y_pred = model_rc.predict(X_test)
    trial_predictions["Power-Law"] = y_pred
    trial_losses["Power-Law"] = model_rc.evaluate(X_test, y_test, metrics=["mae"])[0]

    return trial_losses, trial_predictions

# ----------------------
# --- Main Program -----
# ----------------------
def main():

    bm.set_dt(0.05)
    excel_path = os.path.join(os.getcwd(), 'Losses_Read-in_FixedRL.xlsx')
    sheet_name_median = 'Exp'+args.constraint_set+'_Mackey-Glass_Median'
    sheet_name_iqr = 'Exp'+args.constraint_set+'_Mackey-Glass_IQR'

    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        wb = Workbook()

    ws_m = wb[sheet_name_median] if sheet_name_median in wb.sheetnames else wb.create_sheet(title=sheet_name_median)
    ws_iqr = wb[sheet_name_iqr] if sheet_name_iqr in wb.sheetnames else wb.create_sheet(title=sheet_name_iqr)

    losses = {
        "Uniform": [],
        **{f"Gaussian_sd_{sd}": [] for sd in sd_list},
        "Power-Law": [],
        "Double-Gaussian": [],
        "Laplace": []
    }

    # PRE-INITIALIZE losses and predictions lists for all trials
    for key in losses.keys():
        losses[key] = [[] for _ in range(args.n_trials)]
    predictions = {key: [[] for _ in range(args.n_trials)] for key in losses.keys()}

    start_time = time.time()

    for trial_outer in range(args.n_trials):
        print(f"Outer Trial {trial_outer+1}/{args.n_trials} - Creating fresh model")
        X_train, X_test, y_train, y_test = args.task(n_batch=200, n_states=2)
        model_rc, reservoir_layer = create_base_model(
            (X_train.shape[1], X_train.shape[2]),
            (y_train.shape[1], y_train.shape[2])
        )
        model_serialized = pickle.dumps(model_rc)

        # No need to append here since pre-initialized

        with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
            futures = [executor.submit(run_inner_trial, model_serialized, sd_list, X_train, y_train, X_test, y_test)
                       for _ in range(args.n_inner)]
            for future in concurrent.futures.as_completed(futures):
                trial_result, trial_preds = future.result()
                for key in ["Uniform"] + [f"Gaussian_sd_{sd}" for sd in sd_list] + ["Power-Law", "Laplace"]:
                    losses[key][trial_outer].append(trial_result[key])
                    predictions[key][trial_outer].append(trial_preds[key])

        avg_gauss_losses = {sd: np.mean(losses[f"Gaussian_sd_{sd}"][trial_outer]) for sd in sd_list}
        best_sd = min(avg_gauss_losses, key=avg_gauss_losses.get)
        best_sd_numeric = float(best_sd)
        print(f"Best Gaussian SD for trial {trial_outer+1}: {best_sd_numeric}")

        double_gaussian_inner_losses = []
        double_gaussian_inner_preds = []
        for _ in range(args.n_inner):
            double_weights = create_weights((200, 2), "double_gaussian", dynamic_sd=best_sd_numeric)
            model_rc_dg = pickle.loads(model_serialized)
            model_rc_dg._set_readin_weights(double_weights)
            model_rc_dg.fit(X_train, y_train)
            loss = model_rc_dg.evaluate(X_test, y_test, metrics=["mae"])[0]
            double_gaussian_inner_losses.append(loss)
            y_pred_dg = model_rc_dg.predict(X_test)
            double_gaussian_inner_preds.append(y_pred_dg)

        losses["Double-Gaussian"][trial_outer] = double_gaussian_inner_losses
        predictions["Double-Gaussian"][trial_outer] = double_gaussian_inner_preds

        # --- PLOTTING ---
        '''Please note that The x-axis goes from washout to y_test.shape[1]. 
        y_test.shape[1] is only the length of your output prediction window (n_time_out), not the full sequence of your Mackey-Glass data.
        It is a slicing prediction.'''
        
        print("\nTarget vs Predictions plot..")
        plt.figure(figsize=(12, 6))
        washout = min(200, y_test.shape[1] - 1)
        # Plot true target: these correspond to ground truth output sequence
        plt.plot(range(washout, y_test.shape[1]), y_test[0, washout:, 0], label="Target Signal", color="black", linewidth=2) # it only plots the length of the output prediction window from washout to y_test.shape[1]

        # Prepare method list and colors
        method_list = ["Uniform"] + [f"Gaussian_sd_{sd}" for sd in sd_list] + ["Double-Gaussian", "Laplace", "Power-Law"]
        colors = ['orange', 'green', 'red', 'purple', 'brown', 'pink', 'cyan', 'yellow', 'navy']

        for color, method in zip(colors, method_list):
            preds_for_outer = predictions[method][trial_outer]
            n_inner_preds = len(preds_for_outer)

            if n_inner_preds == 0:
                continue

            if method == "Double-Gaussian":
                y_pred_plot = np.mean(preds_for_outer, axis=0)
            else:
                y_pred_plot = preds_for_outer[0]

            # Determine prediction length
            if y_pred_plot.ndim == 3:
                pred_len = y_pred_plot.shape[1]
                vals = y_pred_plot[0, :pred_len, 0]
            elif y_pred_plot.ndim == 2:
                pred_len = y_pred_plot.shape[0]
                vals = y_pred_plot[:pred_len, 0]
            else:
                pred_len = len(y_pred_plot)
                vals = y_pred_plot[:pred_len]

            # Plot predictions with x-axis shifted to start at washout (aligned with target)
            plt.plot(range(washout, washout + pred_len), vals, label=method, alpha=0.8, color=color)

        plt.title("Target vs Predictions - Mackey-Glass Task")
        plt.xlabel("Time step")
        plt.ylabel("Output")
        handles, labels = plt.gca().get_legend_handles_labels()
        by_label = dict(zip(labels, handles))
        plt.legend(by_label.values(), by_label.keys())
        plt.show()

    # --- Write to Excel ---
    model_columns = {
        "Uniform": 1,
        "Gaussian_sd_0.1": 2,
        "Gaussian_sd_0.25": 3,
        "Gaussian_sd_0.5": 4,
        "Gaussian_sd_0.75": 5,
        "Gaussian_sd_1.0": 6,
        "Double-Gaussian": 7,
        "Laplace": 8,
        "Power-Law": 9
    }

    # Write headers if missing
    for method, col in model_columns.items():
        if ws_m.cell(row=1, column=col).value != method:
            ws_m.cell(row=1, column=col, value=method)
        if ws_iqr.cell(row=1, column=col).value != method:
            ws_iqr.cell(row=1, column=col, value=method)

    # Find first available row
    start_row = 2
    while any(ws_m.cell(row=start_row, column=col).value is not None for col in model_columns.values()):
        start_row += 1

    # Write medians and IQRs
    for trial_outer in range(args.n_trials):
        row_idx = start_row + trial_outer
        for method, col in model_columns.items():
            all_losses = losses[method][trial_outer]

            # Safe flatten
            if all_losses and isinstance(all_losses[0], list):
                all_losses_flat = [item for sublist in all_losses for item in sublist]
            else:
                all_losses_flat = all_losses

            if not all_losses_flat:
                #print(f"[WARNING] No losses recorded for {method}, trial {trial_outer}")
                continue

            median, iqr = compute_median_and_iqr(all_losses_flat)
            ws_m.cell(row=row_idx, column=col, value=median)
            ws_iqr.cell(row=row_idx, column=col, value=iqr)
            print(f"{method} - Median: {median:.6f}, IQR: {iqr:.6f}")

    # Write best Gaussian SD as comment in first cell of column 9
    ws_m.cell(row=1, column=10, value=f"sigma={best_sd_numeric}")

    wb.save(excel_path)
    print(f"Total time: {time.time() - start_time:.2f} sec")

if __name__ == "__main__":
    main()